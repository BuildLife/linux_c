#!/usr/bin/python

import os
import sys
import time
import datetime
#import xlrd
#import xlwt
#import xdrlib
from openpyxl.reader.excel import load_workbook  ### apt-get install python-openpyxl
import threading
import serial as ser
import pandas as pd

freqlist = ["frequency 0 117000000",
			"yes",
			"frequency 1 171000000",
			"yes",
			"frequency 2 255000000",
			"yes",
			"frequency 3 279000000",
			"yes",
			"frequency 4 333000000",
			"yes",
			"frequency 5 387000000",
			"yes",
			"frequency 6 441000000",
			"yes",
			"frequency 7 495000000",
			"yes",
			"frequency 8 549000000",
			"yes",
			"frequency 9 603000000",
			"yes",
			"frequency 10 657000000",
			"yes",
			"frequency 11 711000000",
			"yes",
			"frequency 12 765000000",
			"yes",
			"frequency 13 819000000",
			"yes",
			"frequency 14 861000000",
			"yes",
			"frequency 15 927000000",
			"yes",
			"frequency 16 998000000",
			"yes"]

coeflist = ["coef if 0 1 1000",
			"yes",
			"coef if 1 1 1000",
			"yes",
			"coef if 2 1 1000",
			"yes",
			"coef if 3 1 1000",
			"yes",
			"coef if 4 1 1000",
			"yes",
			"coef if 5 1 1000",
			"yes",
			"coef if 6 1 1000",
			"yes",
			"coef if 7 1 1000",
			"yes",
			"coef if 8 1 1000",
			"yes",
			"coef if 9 1 1000",
			"yes",
			"coef if 10 1 1000",
			"yes",
			"coef if 11 1 1000",
			"yes",
			"coef if 12 1 1000",
			"yes",
			"coef if 13 1 1000",
			"yes",
			"coef if 14 1 1000",
			"yes",
			"coef if 15 1 1000",
			"yes",
			"coef if 16 1 1000",
			"yes",
			"coef if 17 1 1000",
			"yes"]


Tuner1list = ["offset (1) 0 -5",
			"offset (1)	1 -6.07",
			"offset (1)	2 -5.76",
			"offset (1)	3 -5.68",
			"offset (1)	4 -5.4",
			"offset (1)	5 -6.2",
			"offset (1)	6 -6.81",
			"offset (1)	7 -6.4",
			"offset (1)	8 -4.64",
			"offset (1)	9 -5.07",
			"offset (1)	10 -5.54",
			"offset (1)	11 -5.47",
			"offset (1)	12 -6.3",
			"offset (1)	13 -5.88",
			"offset (1)	14 -5.03",
			"offset (1)	15 -5.5",
			"offset (1)	16 -4.35"]


#Enable Factory mode Command
SnmpCommandOpen = ["snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.1.2.1.2.1 s password",
				   "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.1.1.0 i 2",
				   "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.1.1.0 i 2"]

#Disable Factory mode Command
SnmpCommandClose = ["snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.1.1.0 i 0",
				    "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.1.1.0 i 1",
					"snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.1.1.0 i 1"]

#Set Mac address 
SnmpCommandMacList = ["snmpset -v 2c -c sagem 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.1.4.1.2.1 x ",
					  "snmpset -v 2c -c sagem 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.1.4.1.2.2 x ",
					  "snmpset -v 2c -c sagem 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.1.4.1.2.3 x ",
					  "snmpset -v 2c -c sagem 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.1.4.1.2.4 x ",
					  "snmpset -v 2c -c sagem 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.1.4.1.2.5 x "]
#Set Bpi key
SnmpCommandKeyList = ["snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.2.1.0 x 0x",
					  "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.2.2.0 x 0x",
					  "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.2.3.0 x 0x",
					  "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.2.4.0 x 0x",
					  "snmpset -v 2c -c private 192.168.100.1 .1.3.6.1.4.1.4413.2.99.1.1.2.2.2.5.0 x 0x"]

Excelfilename = 'test_excel.xlsx'

CMport = '/dev/ttyUSB0'
CMbaud = 115200

CMFile = '/var/lib/tftpboot'

BroadNUM = sys.argv[1]
BroadNO = sys.argv[2]

print 'Connect to CM.......'
Conn = ser.Serial(CMport, CMbaud, timeout = 0, bytesize=ser.EIGHTBITS, parity=ser.PARITY_NONE, stopbits=ser.STOPBITS_ONE, xonxoff=False, rtscts=False, dsrdtr=False, writeTimeout=0)
if not Conn.isOpen():
	sys.exit()


cMACNum=0
KEYNum=0
cRetnum=0
maclist = []
keylist = []

Cflag = False


### Reading MAC and Key excel file ###
def ReadExcel(f_name, nMac, nKey, retnum):
	print 'Read MAC & KEY excel'
	
	# Read excel file
	xlsx = load_workbook(f_name)

	# Get first table
	Fsheetnames = xlsx.get_sheet_names()
	Fcontent = xlsx.get_sheet_by_name(Fsheetnames[0])
	if not str(Fcontent.title) == 'MAC&KEY':
		print 'Please check excel first page name, must be "MAC&KEY"\n'
		sys.exit()

	# Excel table title
	# row = A	--> Number
	# 		B C --> STB MAC
	#		D 	--> Broad NO.
	#		E 	--> Date
	#		F 	--> CM MAC
	# 		H 	--> cmBpiPublicKey (281)
	#		I 	--> cmBpiPrivateKey (1271)
	#		J 	--> cmBpiPlusRootPublicKey (541)
	#		K 	--> cmBpiPlusCmCertificate (1571)
	#		L 	--> cmBpiPlusCaCertificate (2475)
	
	Keyrowlist = ['H','I','J','K','L']
	
	# Find module
	for num in range(2,300):
		if Fcontent.cell('A' + str(num)).value == int(BroadNUM):
			print 'Find out it'
			retnum = num

	if retnum == 0:
		print 'NO '+ BroadNUM +' Module, please check it again' 
		os._exit(0)

	# Get table values
	if Fcontent.cell('D1').value == "Broad No.":
		if Fcontent.cell('F1').value == "CM MAC":
			# Set MAC address to maclist
			for x in range(retnum,retnum+5):
				maclist.insert(nMac,Fcontent.cell('F'+ str(x)).value)
				nMac+=1
			# Set Key to keylist
			for y in Keyrowlist:
				keylist.insert(nKey,Fcontent.cell(y + str(retnum)).value)
				nKey+=1
		else:
			print 'No " CM MAC " title'
	else:
		print 'No " Broad NO." title'

	# Set new Broad NO. and time
	Fcontent.cell('D' + str(retnum)).value = str(BroadNO)
	Fcontent.cell('E' + str(retnum)).value = str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

	#Fcontent.merge_cells('A' + str(retnum) + ':' + 'A' + str(retnum+5))
	xlsx.save(f_name)

	return retnum

'''# write broad number to excel
#def WriteExcel(f_name, wBroadNUM, wBroadNO):
#	print 'Write Broad NO: '
'''		

### Set frequnecy & coef & tuner1 in Console ###
def WriteFreq():
	print 'Write frequnecy & coef & Tuner1 in Console'
	#Conn.write("cd /"+"\n")
	#Conn.write("/doc/sc"+"\n")
	#Conn.flush()
	while True:
		if Cflag:
			time.sleep(5)
			Conn.write("cd non/ds" + "\n")
			#print data.find("CM/NonVol/Ds Cal NonVol>")
			#if data.find("CM/NonVol/Ds Cal NonVol>") > 0:
			time.sleep(5)
			if Cflag:
				print 'Enter in DS Cal NonVol folder'
				Conn.write("num_freq 17" + "\n")
				time.sleep(0.5)
				Conn.write("yes" + "\n")
				time.sleep(0.5)
				Conn.write("write" + "\n")
				time.sleep(0.5)
				Conn.write("yes" + "\n")
				time.sleep(0.5)
				for freq in freqlist:
					Conn.write(freq + "\n")
					time.sleep(0.5)
				Conn.write("write" + "\n")
				Conn.write("yes" + "\n")
				print 'write freq done'
				for coef in coeflist:
					Conn.write(coef + "\n")
					time.sleep(0.5)
				Conn.write("write" + "\n")
				Conn.write("yes" + "\n")
				print 'write coef done'
				for tuner1 in Tuner1list:
					Conn.write(tuner1 + "\n")
					time.sleep(0.5)
				Conn.write("write" + "\n")
				Conn.write("yes" + "\n")
				print 'write Tuner1 done'
				Conn.flush()
				break
		
### Set Mac Address in Console ###
def WriteMac(nMac, nKey):
	print 'Start to Write MAC in Console'
	Conn.write("cd /"+"\n")
	#if buf.find("CM> ") == 0:
	if maclist and keylist:
		Conn.write("cd non/hal" + "\n")
		time.sleep(1)
		#if buf.find("CM/NonVol/HalIf NonVol>") == 0:
		conmac = ''
		for mac in maclist:
			for c in range(0,12,2):
				if c == 10:
					conmac = conmac + mac[c:c+2]
					break
				else:
					conmac = conmac + mac[c:c+2] + ':'
			Conn.write("mac_address "+ str((maclist.index(mac))+1) + " " + conmac + "\n")
			conmac = ''
			Conn.write("write" + "\n")
			time.sleep(0.5)
		Conn.write("write" + "\n")
		#Conn.write("yes" + "\n")
		print 'Set mac address down'
	else:
		print 'No key and mac can set in CM'
		print 'Please check excel content'
		#sys.exit()
		os._exit(0)


### Snmp Command for set MAC & Key ###
def SnmpCmd():
	print 'Snmp Command to Set MAC and Key'
	for snmpcmdopen in SnmpCommandOpen:
		os.system(snmpcmdopen)
	time.sleep(1)
#	for snmpcmdmac in SnmpCommandMacList:
#		os.system(snmpcmdmac + '"0x' + str(maclist[SnmpCommandMacList.index(snmpcmdmac)]) + '"')
#		time.sleep(1)
	for snmpcmdkey in SnmpCommandKeyList:
		os.system(snmpcmdkey + str(keylist[SnmpCommandKeyList.index(snmpcmdkey)]))
		time.sleep(1)
	for snmpcmdclose in SnmpCommandClose:
		os.system(snmpcmdclose)
	print 'Snmp Command set down............'




### Reading Console log ###
data=''
def Read_thread(Conn):
	global Cflag
	print 'Start to read....'
	while True:
		try:
			data = Conn.read(1000)
			#data = Conn.readline()
			#data = Conn.readall()
			#print data.find('CM>')
			if data.find('CM>') > 0:
				Cflag = True
			elif data.find('Scanning DS Channel') > 0 or data.find('Checking plant power') > 0:
				Conn.write("/doc/sc" + "\n")
		#	print 'cflag = ' + str(Cflag)
			if data != '':
				print data
			time.sleep(0.02)
		except KeyboardInterrupt:
			print 'exiting'
			break
	Conn.flush()
	Conn.close()
	os_.exit(0)
		

if __name__=='__main__':
	# Reading Console Thread
	thread = threading.Thread(target=Read_thread, args=(Conn,))
	thread.start()

	# Reading Key and MAC excel
	cRetnum = ReadExcel(Excelfilename, cMACNum, KEYNum, cRetnum)
	Conn.write("\n")
	Conn.write("cd /" + "\n")
	while True:
		try:
			if Cflag:
				WriteFreq()
				time.sleep(1)
				Conn.write("cd /" + "\n")
				time.sleep(5)
				WriteMac(cRetnum,0)
				print 'Enter in snmp command......'
				SnmpCmd()
				print '*******************************************'
				print '**** Finish set MAC & KEY, Good Bye... ****'
				print '*******************************************'
				Conn.flush()
				Conn.close()
				break
		except KeyboardInterrupt:
			print 'exiting'
			break
	os._exit(0)


'''
def UpgradeFirmware(Conn):
	print 'Start to upgrade firmware'
	Conn.write("/res" + "\n")
	print 'restart'
	time.sleep(7)
	Conn.write(CMBroadComPwd + "\n")
	if buf.find("  within 2 seconds or take default...") == -1:
	  	print 'Enter in upgrade mode'
		Conn.write("p")


UpgradeFirmware(Conn)
'''







